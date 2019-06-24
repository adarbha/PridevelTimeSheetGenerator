import calendar
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

FULL_WORK_HRS = 8.0

class TimeSheetGenerator(object):
    '''Need a config_json dictionary from the calling object. All file generation specific parameters in config_json'''
    def __init__(self,config_json):
        self.config_json = config_json
        self.wb = load_workbook(filename=config_json["sample_file_location"])
        # Only the first sheet would be used - so hard codeing this
        self.sheet = self.wb[self.wb.sheetnames[0]] 
        self.dates_to_cell = self.map_dates_to_cells()
        self.dates = []
        self.generate_dates()

    def map_dates_to_cells(self):
        '''Dates 1 - 15 is range B13 to P13 and 16 - 31 is range B14 to Q14. Generate a dict with that info.
        Sheet object is the input. Returns a dictionary - date:cell_object'''
        dates_to_cell = {}
        dates_to_cell = dates_to_cell.fromkeys(range(1,32)) #Keys are dates from 1,31
        #Cell ranges for work hour cells are hard-coded. This never changes
        cell_range_B13_P13 = self.sheet['B13:P13'][0]
        cell_range_B14_Q14 = self.sheet['B14:Q14'][0] 

        for key,value in dates_to_cell.items():
            if (key > 0) and (key < 16):
                dates_to_cell[key] = cell_range_B13_P13[key-1]
            else:
                dates_to_cell[key] = cell_range_B14_Q14[key - 16]
    
        return dates_to_cell

    def reset_all_cells(self):
        '''Resets all cells in range to zero'''
        for key in self.dates_to_cell.keys():
            self.dates_to_cell[key].value = 0
        
    def fill_cells(self):
        '''fill_cells based on dates. Assuming all dates have 8 full work hours'''
        for date in self.dates:
            self.dates_to_cell[date.day].value = FULL_WORK_HRS
            

    def generate_dates(self):
        '''Generate dates based on date ranges from config_json'''
        ##TODO Holidays

        self.dates = pd.bdate_range(start=self.config_json['start_date'],end=self.config_json['end_date'],holidays=[],freq='C').tolist()



    def fill_fillers(self):
        '''This is to fill meta-data in various cells. Al the cells here are harcoded'''
        #O4 is the month for which time sheet is generated
        self.sheet['O4'].value = self.dates[0].month_name()
        #P4 is the year for which time sheet is generated
        self.sheet['P4'].value = self.dates[0].year

        dates = [x.day for x in self.dates]
        #O5 x is if dates between 1 to 15
        #O6 x if dates between 16 to 31
        for d in dates:
            if d in range(1,16):
                self.sheet['O5'].value = 'x'
            if d in range(16,32):
                self.sheet['O6'].value = 'x'
        
        #Aggregations - adding formulae
        #R13 and R14 for row sums
        self.sheet['R13'].value = "=SUM(B13:Q13)"
        self.sheet['R14'].value = "=SUM(B14:Q14)"


        #R36 is todays date
        now = datetime.now()
        self.sheet['R36'].value = "{}-{}-{}".format(now.day,calendar.month_abbr[now.month],now.year)
        #Insert image
        img = Image("pridevel_image.png")
        self.sheet.add_image(img,'A1')
        

    def save_file(self):
        self.reset_all_cells()
        self.fill_cells()
        self.fill_fillers()
        self.wb.save(filename=self.config_json['file_to_generate'])


if __name__ == "__main__":
    test_config = {}
    test_config['sample_file_location'] = 'sample_pridevel.xlsx'
    test_config['file_to_generate'] = 'test.xlsx'
    test_config['start_date'] = "02-11-2019"
    test_config['end_date'] = "02-14-2019"

    #Generate a Timesheetgenerator object - This is used for testing only
    TG = TimeSheetGenerator(test_config)
    TG.reset_all_cells()
    TG.save_file()
